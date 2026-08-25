"""
run_simplex.py
===============
Gera um relatório Excel (.xlsx) a partir dos resultados de um grid search do
algoritmo Simplex, comparando o custo obtido em cada execução com o valor
ótimo de referência da biblioteca Netlib.

Uso:
    python run_simplex.py
    python run_simplex.py --csv resultados_grid_search.csv --output Relatorio.xlsx
    python run_simplex.py --tolerance 1e-5

Critério de cor (aplicado em todas as abas):
    Verde   -> execução com SUCCESS e custo bate com o valor ótimo (dentro da tolerância)
    Amarelo -> execução com SUCCESS mas custo diferente do valor ótimo
    Vermelho -> qualquer status diferente de SUCCESS (erro, timeout, etc.)

Nas abas de resumo, "sucesso" conta apenas quando o resultado é ótimo (verde).
Um teste que apenas "rodou sem erro" mas com valor errado não é contado como
sucesso -- é tratado como parcial (amarelo).
"""

import argparse
import logging
import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# --------------------------------------------------------------------------- #
# Configuração
# --------------------------------------------------------------------------- #

DEFAULT_CSV_FILE = "resultados_grid_search.csv"
DEFAULT_EXCEL_FILE = "Resultados_Simplex_Parametros.xlsx"
DEFAULT_TOLERANCE = 1e-4  # tolerância relativa para considerar o custo "ótimo"

REQUIRED_COLUMNS = [
    "Run_ID", "Instance", "Status", "Cost", "Time_Seconds",
    "Epsilon", "Preprocess", "Seed", "Refactor_Period",
]
PARAM_COLS = ["Preprocess", "Seed", "Epsilon_Num", "Refactor_Period"]

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Sucesso exato
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Parcial / valor errado
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Erro / sem acertos

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
BODY_FONT = Font(name="Arial")

# Formatos numéricos por nome de coluna (aplicados quando a coluna existe na aba)
NUMBER_FORMATS = {
    "Cost_Num": "#,##0.0000",
    "Optimal_Ref": "#,##0.0000",
    "Abs_Diff": "#,##0.0000",
    "Rel_Diff": "0.0000%",
    "Time_Seconds": "0.000",
    "Tempo_Medio_s": "0.000",
    "Taxa_Acerto_Otimo": "0.0%",
    "Epsilon_Num": "0.0E+00",
}

# --------------------------------------------------------------------------- #
# Referência Netlib
# --------------------------------------------------------------------------- #

NETLIB_REFERENCE_DATA = """
25FV47      822   1571    11127      70477        5.5018458883E+03
80BAU3B    2263   9799    29063     298952  B     9.8723216072E+05
ADLITTLE     57     97      465       3690        2.2549496316E+05
AFIRO        28     32       88        794       -4.6475314286E+02
AGG         489    163     2541      21865       -3.5991767287E+07
AGG2        517    302     4515      32552       -2.0239252356E+07
AGG3        517    302     4531      32570        1.0312115935E+07
BANDM       306    472     2659      19460       -1.5862801845E+02
BEACONFD    174    262     3476      17475        3.3592485807E+04
BLEND        75     83      521       3227       -3.0812149846E+01
BNL1        644   1175     6129      42473        1.9776292856E+03
BNL2       2325   3489    16124     127145        1.8112365404E+03
BOEING1     351    384     3865      25315  BR   -3.3521356751E+02
BOEING2     167    143     1339       8761  BR   -3.1501872802E+02
BORE3D      234    315     1525      13160  B     1.3730803942E+03
BRANDY      221    249     2150      14028        1.5185098965E+03
CAPRI       272    353     1786      15267  B     2.6900129138E+03
CYCLE      1904   2857    21322     166648  B    -5.2263930249E+00
CZPROB      930   3523    14173      92202  B     2.1851966989E+06
D2Q06C     2172   5167    35674     258038        1.2278423615E+05
D6CUBE      416   6184    43888     167633  B     3.1549166667E+02
DEGEN2      445    534     4449      24657       -1.4351780000E+03
DEGEN3     1504   1818    26230     130252       -9.8729400000E+02
DFL001     6072  12230    41873     353192  B     1.12664E+07
E226        224    282     2767      17749       -1.8751929066E+01
ETAMACRO    401    688     2489      21915  B    -7.5571521774E+02
FFFFF800    525    854     6235      39637        5.5567961165E+05
FINNIS      498    614     2714      23847  B     1.7279096547E+05
FIT1D        25   1026    14430      51734  B    -9.1463780924E+03
FIT1P       628   1677    10894      65116  B     9.1463780924E+03
FIT2D        26  10500   138018     482330  B    -6.8464293294E+04
FIT2P      3001  13525    60784     439794  B     6.8464293232E+04
FORPLAN     162    421     4916      25100  BR   -6.6421873953E+02
GANGES     1310   1681     7021      60191  B    -1.0958636356E+05
GFRD-PNC    617   1092     3467      24476  B     6.9022359995E+06
GREENBEA   2393   5405    31499     235711  B    -7.2462405908E+07
GREENBEB   2393   5405    31499     235739  B    -4.3021476065E+06
GROW15      301    645     5665      35041  B    -1.0687094129E+08
GROW22      441    946     8318      50789  B    -1.6083433648E+08
GROW7       141    301     2633      17043  B    -4.7787811815E+07
ISRAEL      175    142     2358      12109       -8.9664482186E+05
KB2          44     41      291       2526  B    -1.7499001299E+03
LOTFI       154    308     1086       6718       -2.5264706062E+01
MAROS       847   1443    10006      65906  B    -5.8063743701E+04
MAROS-R7   3137   9408   151120    4812587        1.4971851665E+06
MODSZK1     688   1620     4158      40908  B     3.2061972906E+02
NESM        663   2923    13988     117828  BR    1.4076073035E+07
PEROLD      626   1376     6026      47486  B    -9.3807580773E+03
PILOT      1442   3652    43220     278593  B    -5.5740430007E+02
PILOT.JA    941   1988    14706      97258  B    -6.1131344111E+03
PILOT.WE    723   2789     9218      79972  B    -2.7201027439E+06
PILOT4      411   1000     5145      40936  B    -2.5811392641E+03
PILOT87    2031   4883    73804     514192  B     3.0171072827E+02
PILOTNOV    976   2172    13129      89779  B    -4.4972761882E+03
QAP8        913   1632     8304                   2.0350000000E+02
QAP12      3193   8856    44244                   5.2289435056E+02
QAP15      6331  22275   110700                   1.0409940410E+03
RECIPE       92    180      752       6210  B    -2.6661600000E+02
SC105       106    103      281       3307       -5.2202061212E+01
SC205       206    203      552       6380       -5.2202061212E+01
SC50A        51     48      131       1615       -6.4575077059E+01
SC50B        51     48      119       1567       -7.0000000000E+01
SCAGR25     472    500     2029      17406       -1.4753433061E+07
SCAGR7      130    140      553       4953       -2.3313892548E+06
SCFXM1      331    457     2612      19078        1.8416759028E+04
SCFXM2      661    914     5229      37079        3.6660261565E+04
SCFXM3      991   1371     7846      53828        5.4901254550E+04
SCORPION    389    358     1708      12186        1.8781248227E+03
SCRS8       491   1169     4029      36760        9.0429998619E+02
SCSD1        78    760     3148      17852        8.6666666743E+00
SCSD6       148   1350     5666      32161        5.0500000078E+01
SCSD8       398   2750    11334      65888        9.0499999993E+02
SCTAP1      301    480     2052      14970        1.4122500000E+03
SCTAP2     1091   1880     8124      57479        1.7248071429E+03
SCTAP3     1481   2480    10734      78688        1.4240000000E+03
SEBA        516   1028     4874      38627  BR    1.5711600000E+04
SHARE1B     118    225     1182       8380       -7.6589318579E+04
SHARE2B      97     79      730       4795       -4.1573224074E+02
SHELL       537   1775     4900      38049  B     1.2088253460E+09
SHIP04L     403   2118     8450      57203        1.7933245380E+06
SHIP04S     403   1458     5810      41257        1.7987147004E+06
SHIP08L     779   4283    17085     117083        1.9090552114E+06
SHIP08S     779   2387     9501      70093        1.9200982105E+06
SHIP12L    1152   5427    21597     146753        1.4701879193E+06
SHIP12S    1152   2763    10941      82527        1.4892361344E+06
SIERRA     1228   2036     9252      76627  B     1.5394362184E+07
STAIR       357    467     3857      27405  B    -2.5126695119E+02
STANDATA    360   1075     3038      26135  B     1.2576995000E+03
STANDMPS    468   1075     3686      29839  B     1.4060175000E+03
STOCFOR1    118    111      474       4247       -4.1131976219E+04
STOCFOR2   2158   2031     9492      79845       -3.9024408538E+04
STOCFOR3  16676  15695    74004                  -3.9976661576E+04
TRUSS      1001   8806    36642                   4.5881584719E+05
TUFF        334    587     4523      29439  B     2.9214776509E-01
VTP.BASE    199    203      914       8175  B     1.2983146246E+05
WOOD1P      245   2594    70216     328905        1.4429024116E+00
WOODW      1099   8405    37478     240063        1.3044763331E+00
"""


def parse_netlib_reference(raw_text: str) -> dict:
    """Converte o bloco de texto da Netlib em {nome_instancia: valor_otimo}."""
    ref_map = {}
    for line in raw_text.strip().split("\n"):
        tokens = line.strip().split()
        if not tokens:
            continue
        name = tokens[0].upper()
        try:
            ref_map[name] = float(tokens[-1].replace("**", ""))
        except ValueError:
            logger.warning("Não foi possível interpretar o valor ótimo da linha: %r", line)
            ref_map[name] = None
    return ref_map


def normalize_inst_name(name) -> str:
    """Normaliza o nome da instância (remove extensão, corrige VTPBASE)."""
    clean = str(name).upper().strip()
    clean = re.sub(r"\.(MPS|QPS|SIF)$", "", clean)
    return "VTP.BASE" if clean == "VTPBASE" else clean


# --------------------------------------------------------------------------- #
# Formatação da planilha
# --------------------------------------------------------------------------- #

def style_header(ws: Worksheet) -> None:
    """Aplica negrito/cor no cabeçalho e congela a primeira linha."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    if ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions


def apply_number_formats(ws: Worksheet) -> None:
    """Aplica formato numérico às colunas conhecidas, pela posição do cabeçalho."""
    headers = [cell.value for cell in ws[1]]
    for col_name, fmt in NUMBER_FORMATS.items():
        if col_name not in headers:
            continue
        col_idx = headers.index(col_name)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[col_idx].number_format = fmt


def autofit_columns(ws: Worksheet, min_width: int = 10, max_width: int = 40) -> None:
    """Ajusta a largura das colunas ao conteúdo (aproximado)."""
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


def set_body_font(ws: Worksheet) -> None:
    """Garante fonte consistente (Arial) no corpo da tabela."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT


def apply_detail_colors(ws: Worksheet) -> None:
    """Aplica cores para tabelas detalhadas (por instância e grid search).

    Verde   : Status == SUCCESS e Is_Optimal == True
    Amarelo : Status == SUCCESS e Is_Optimal == False
    Vermelho: qualquer outro Status (erro, timeout, etc.)
    """
    headers = [cell.value for cell in ws[1]]
    if "Status" not in headers or "Is_Optimal" not in headers:
        return

    status_idx = headers.index("Status")
    optimal_idx = headers.index("Is_Optimal")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status_val = str(row[status_idx].value).strip().upper()
        is_optimal = bool(row[optimal_idx].value)

        if status_val == "SUCCESS" and is_optimal:
            fill = GREEN_FILL
        elif status_val == "SUCCESS" and not is_optimal:
            fill = YELLOW_FILL
        else:
            fill = RED_FILL

        for cell in row:
            cell.fill = fill


def apply_summary_colors(ws: Worksheet) -> None:
    """Aplica cores para tabelas resumidas de agrupamento de parâmetros.

    "Sucesso" aqui é sempre sinônimo de "ótimo" (verde) -- uma execução que
    apenas rodou sem erro, mas com valor diferente do ótimo, não conta como
    sucesso e cai em amarelo (parcial).

    Verde   : Acertos_Otimos == Total_Testes (100% de acerto exato)
    Amarelo : 0 < Acertos_Otimos < Total_Testes (acerto parcial)
    Vermelho: Acertos_Otimos == 0 (nenhum acerto exato)
    """
    headers = [cell.value for cell in ws[1]]
    if "Acertos_Otimos" not in headers or "Total_Testes" not in headers:
        return

    otimos_idx = headers.index("Acertos_Otimos")
    total_idx = headers.index("Total_Testes")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        try:
            otimos = float(row[otimos_idx].value or 0)
            total = float(row[total_idx].value or 0)
        except (ValueError, TypeError):
            continue

        if total <= 0:
            continue
        elif otimos == total:
            fill = GREEN_FILL
        elif otimos > 0:
            fill = YELLOW_FILL
        else:
            fill = RED_FILL

        for cell in row:
            cell.fill = fill


def finalize_sheet(ws: Worksheet, color_fn) -> None:
    """Aplica cor, formato numérico, fonte e layout padrão a uma aba."""
    color_fn(ws)
    style_header(ws)
    set_body_font(ws)
    apply_number_formats(ws)
    autofit_columns(ws)


# --------------------------------------------------------------------------- #
# Processamento dos dados
# --------------------------------------------------------------------------- #

def load_results(csv_file: str) -> pd.DataFrame:
    """Lê e valida o CSV de resultados do grid search."""
    path = Path(csv_file)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo de resultados não encontrado: '{csv_file}'")

    df = pd.read_csv(path)

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            f"O CSV '{csv_file}' está sem as colunas obrigatórias: {missing}. "
            f"Colunas encontradas: {list(df.columns)}"
        )

    if df.empty:
        raise ValueError(f"O CSV '{csv_file}' não contém nenhuma linha de dados.")

    return df


def enrich_results(df: pd.DataFrame, netlib_ref: dict, tolerance: float) -> pd.DataFrame:
    """Adiciona colunas derivadas: referência ótima, diferenças e flags de sucesso."""
    df = df.copy()

    df["Norm_Instance"] = df["Instance"].apply(normalize_inst_name)
    df["Optimal_Ref"] = df["Norm_Instance"].map(netlib_ref)

    unmatched = sorted(set(df.loc[df["Optimal_Ref"].isna(), "Norm_Instance"]))
    if unmatched:
        logger.warning(
            "%d instância(s) sem valor de referência na Netlib (não entram no cálculo de acerto): %s",
            len(unmatched), ", ".join(unmatched),
        )

    df["Cost_Num"] = pd.to_numeric(df["Cost"], errors="coerce")
    df["Time_Seconds"] = pd.to_numeric(df["Time_Seconds"], errors="coerce")
    df["Epsilon_Num"] = pd.to_numeric(df["Epsilon"], errors="coerce")

    df["Abs_Diff"] = (df["Cost_Num"] - df["Optimal_Ref"]).abs()
    df["Rel_Diff"] = np.where(
        df["Optimal_Ref"].abs() > 1e-9,
        df["Abs_Diff"] / df["Optimal_Ref"].abs(),
        df["Abs_Diff"],
    )

    df["Is_Success"] = df["Status"].astype(str).str.strip().str.upper() == "SUCCESS"
    df["Is_Optimal"] = df["Is_Success"] & (df["Rel_Diff"] < tolerance)

    return df


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Aba 'Resumo Executivo': uma linha por combinação de parâmetros."""
    summary = df.groupby(PARAM_COLS, as_index=False).agg(
        Total_Testes=("Run_ID", "count"),
        Acertos_Otimos=("Is_Optimal", "sum"),
        Execucoes_Sem_Erro=("Is_Success", "sum"),
        Tempo_Medio_s=("Time_Seconds", "mean"),
    )
    summary["Taxa_Acerto_Otimo"] = summary["Acertos_Otimos"] / summary["Total_Testes"]
    return summary.sort_values(
        by=["Acertos_Otimos", "Taxa_Acerto_Otimo", "Tempo_Medio_s"],
        ascending=[False, False, True],
    )


def build_best_per_instance(df: pd.DataFrame) -> pd.DataFrame:
    """Aba 'Resultados por Instancia': melhor execução de cada instância."""
    best = (
        df.sort_values(
            by=["Instance", "Is_Optimal", "Is_Success", "Time_Seconds"],
            ascending=[True, False, False, True],
        )
        .groupby("Instance", as_index=False)
        .first()
    )
    cols = [
        "Instance", "Status", "Cost_Num", "Optimal_Ref", "Rel_Diff", "Is_Optimal",
        "Time_Seconds", "Preprocess", "Seed", "Epsilon_Num", "Refactor_Period",
    ]
    return best[cols]


def build_param_comparison(df: pd.DataFrame) -> pd.DataFrame:
    """Aba 'Comparacao de Parametros': impacto de cada parâmetro isoladamente."""
    summaries = []
    for param in PARAM_COLS:
        p_df = df.groupby(param, as_index=False).agg(
            Total_Testes=("Run_ID", "count"),
            Acertos_Otimos=("Is_Optimal", "sum"),
            Execucoes_Sem_Erro=("Is_Success", "sum"),
            Tempo_Medio_s=("Time_Seconds", "mean"),
        )
        p_df.insert(0, "Parametro", param)
        p_df.rename(columns={param: "Opcao_Valor"}, inplace=True)
        p_df["Taxa_Acerto_Otimo"] = p_df["Acertos_Otimos"] / p_df["Total_Testes"]
        p_df = p_df.sort_values(by=["Acertos_Otimos", "Tempo_Medio_s"], ascending=[False, True])
        summaries.append(p_df)

    return pd.concat(summaries, ignore_index=True)


# --------------------------------------------------------------------------- #
# Orquestração
# --------------------------------------------------------------------------- #

def generate_excel_report(
    csv_file: str = DEFAULT_CSV_FILE,
    excel_file: str = DEFAULT_EXCEL_FILE,
    tolerance: float = DEFAULT_TOLERANCE,
) -> None:
    df = load_results(csv_file)
    netlib_ref = parse_netlib_reference(NETLIB_REFERENCE_DATA)
    df = enrich_results(df, netlib_ref, tolerance)

    sheets = {
        "Resumo Executivo": (build_summary(df), apply_summary_colors),
        "Resultados por Instancia": (build_best_per_instance(df), apply_detail_colors),
        "Comparacao de Parametros": (build_param_comparison(df), apply_summary_colors),
        "Resultados Grid Search": (df, apply_detail_colors),
    }

    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        for sheet_name, (sheet_df, _) in sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

        for sheet_name, (_, color_fn) in sheets.items():
            finalize_sheet(writer.sheets[sheet_name], color_fn)

    n_optimal = int(df["Is_Optimal"].sum())
    n_success = int(df["Is_Success"].sum())
    logger.info(
        "Relatório gerado com sucesso: '%s' (%d execuções | %d ótimas | %d rodaram sem erro)",
        excel_file, len(df), n_optimal, n_success,
    )


def parse_args(argv=None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Gera relatório Excel do grid search do Simplex.")
    parser.add_argument("--csv", dest="csv_file", default=DEFAULT_CSV_FILE,
                         help=f"Caminho do CSV de entrada (padrão: {DEFAULT_CSV_FILE})")
    parser.add_argument("--output", dest="excel_file", default=DEFAULT_EXCEL_FILE,
                         help=f"Caminho do Excel de saída (padrão: {DEFAULT_EXCEL_FILE})")
    parser.add_argument("--tolerance", type=float, default=DEFAULT_TOLERANCE,
                         help=f"Tolerância relativa para considerar o custo ótimo (padrão: {DEFAULT_TOLERANCE})")
    return parser.parse_args(argv)


def main(argv=None) -> int:
    args = parse_args(argv)
    try:
        generate_excel_report(args.csv_file, args.excel_file, args.tolerance)
    except (FileNotFoundError, ValueError) as e:
        logger.error(str(e))
        return 1
    except Exception:
        logger.exception("Falha inesperada ao gerar o relatório.")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
